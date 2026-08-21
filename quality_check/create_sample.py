# -*- coding: utf-8 -*-
"""生成示例输入 Excel，包含各种测试场景。

运行：python -m quality_check.create_sample
生成：../data/qc_input.xlsx

Excel 格式：
  第1行：合并表头"基本信息"（跨26列）+"数管反馈"（跨3列）
  第2行：29列列名
  第3行+：数据

样例说明：
  - 前 22 条为原有样例（含故意造错与应通过）
  - 后 10 条为新增"应通过"样例，覆盖编码/文本/数值/日期时间/代码枚举等类型
"""

import os

import pandas as pd
from openpyxl import load_workbook


# 申请单全部 29 列的默认空值，保证生成格式与真实申请单一致
_EMPTY = {
    "贴源接口(必填)": "",
    "贴源接口文件中文名(必填)": "",
    "贴源物理字段名(必填)": "",
    "贴源物理字段中文名(必填)": "",
    "贴源物理字段类型/长度(必填)": "",
    "贴源接口联系人姓名(必填)": "",
    "贴源接口联系人邮箱(必填)": "",
    "申请单号(必填)": "",
    "子单编号(必填)": "",
    "中文表名(必填)": "",
    "是否新建表(必填)": "",
    "表业务含义(选填)": "",
    "Schema名称(必填)": "",
    "表英文名(选填)": "",
    "中文字段名(必填)": "",
    "域类型(必填)": "",
    "数据示例(必填)": "",
    "字段所属类型(必填)": "",
    "度量单位(选填)": "",
    "是否主键(选填)": "",
    "更新频次(选填)": "",
    "是否枚举(必填)": "",
    "枚举值(选填)": "",
    "更新标准编号(选填)": "",
    "业务定义(必填)": "",
    "业务口径(选填)": "",
    "格式化枚举值": "",
    "检查结果": "",
    "说明": "",
}


def make_row(table="", field="", domain="", example="", ftype="",
             is_enum="", meaning="", enum_vals=""):
    """构造一行申请单数据，未指定的列自动填空。

    Args:
        table: 中文表名
        field: 中文字段名
        domain: 域类型
        example: 数据示例
        ftype: 字段所属类型
        is_enum: 是否枚举
        meaning: 业务定义
        enum_vals: 枚举值
    """
    row = dict(_EMPTY)
    row["中文表名(必填)"] = table
    row["中文字段名(必填)"] = field
    row["域类型(必填)"] = domain
    row["数据示例(必填)"] = example
    row["字段所属类型(必填)"] = ftype
    row["是否枚举(必填)"] = is_enum
    row["业务定义(必填)"] = meaning
    row["枚举值(选填)"] = enum_vals
    return row


def main():
    data = [
        # ==================== 原有样例（22条） ====================

        # --- 客户信息表 ---
        # 应通过：见名知意
        make_row(table="客户信息表", field="客户名称", domain="anc..(100)",
                 example="张三", ftype="文本类", is_enum="否", meaning="客户名称"),
        # 应通过：见名知意（DATE 示例应为纯日期，此处造错：带时分秒）
        make_row(table="客户信息表", field="开户日期", domain="DATE",
                 example="2000-01-01 00:00:00", ftype="日期时间类", is_enum="否",
                 meaning="开户的日期"),
        # 故意造错：域类型 VARCHAR(100) 非标准格式 + 是否枚举为空
        make_row(table="客户信息表", field="手机号码", domain="VARCHAR(100)",
                 example="11111111111", ftype="文本类", meaning="客户手机号"),
        # 故意造错：域类型与数据示例缺失
        make_row(table="客户信息表", field="客户状态", ftype="代码枚举类",
                 is_enum="是", enum_vals="01-正常;02-冻结;03-注销",
                 meaning="标识客户在系统中的当前状态，包括正常、冻结、注销等"),
        # 故意造错：i(19,2) 小数位数超限（11111.1111 小数4位>2）
        make_row(table="客户信息表", field="交易金额", domain="i(19,2)",
                 example="11111.1111", ftype="数值类", is_enum="否",
                 meaning="记录一笔交易的实际发生金额，单位为元"),
        # 故意造错：业务含义与字段名完全复制
        make_row(table="客户信息表", field="第二结论类型", domain="an..(2)",
                 example="通过", ftype="代码枚举类", is_enum="是",
                 enum_vals="01-通过;02-拒绝", meaning="第二结论类型"),
        # 故意造错：业务含义简单加词（"的"）
        make_row(table="客户信息表", field="CDA规则结果", domain="an..(2)",
                 example="01", ftype="代码枚举类", is_enum="是",
                 enum_vals="01-命中;02-未命中", meaning="CDA的规则结果"),

        # --- 客户信息明细表 ---
        # 故意造错：业务含义简单加词（"字段"）+ 同表重复
        make_row(table="客户信息明细表", field="CDA规则结果", domain="an..(2)",
                 example="01", ftype="代码枚举类", is_enum="是",
                 meaning="CDA规则结果字段"),
        # 故意造错：数据示例缺失 + 同表重复
        make_row(table="客户信息明细表", field="转案欺诈类型", domain="an..(2)",
                 ftype="代码枚举类", is_enum="是", enum_vals="01-欺诈;02-正常",
                 meaning="指转案欺诈的不同类别"),
        # 应通过：业务含义有实质描述（解释CDA）——虽同表重名但字段语义有效
        make_row(table="客户信息明细表", field="CDA规则结果", domain="an..(2)",
                 example="01", ftype="代码枚举类", is_enum="是",
                 enum_vals="01-命中;02-未命中",
                 meaning="信用风险评估系统CDA引擎生成的规则判定结果，用于自动审批决策"),
        # 应通过：业务含义有实质描述（解释转案欺诈）
        make_row(table="客户信息明细表", field="转案欺诈类型", domain="an..(2)",
                 example="01", ftype="代码枚举类", is_enum="是",
                 meaning="标记案件转案过程中涉及的欺诈行为分类，由风控系统自动识别"),
        # 故意造错：中文字段名缺失
        make_row(table="客户信息明细表", domain="an!(10)", ftype="文本类",
                 is_enum="否", meaning="记录客户的电子邮箱地址"),
        # 故意造错：字段所属类型缺失
        make_row(table="客户信息明细表", field="证件类型", domain="an..(2)",
                 example="01", is_enum="是", enum_vals="身份证;护照;军官证",
                 meaning="标识客户使用的证件类型"),
        # 故意造错：域类型"i"格式不合法 + 业务定义缺失 + 是否枚举应为否
        make_row(table="客户信息明细表", field="账户余额", domain="i",
                 example="1111.11", ftype="数值类", is_enum="是"),
        # 故意造错：数据示例长度超限（an..(5) 示例9位）
        make_row(table="客户信息明细表", field="银行卡号", domain="an..(5)",
                 example="111111111", ftype="编码类", is_enum="否",
                 meaning="记录客户绑定的银行卡卡号，用于资金划转"),
        # 故意造错：字段所属类型"布尔型"不合法
        make_row(table="客户信息明细表", field="是否VIP", domain="an..(1)",
                 example="是", ftype="布尔型", is_enum="否",
                 meaning="标识客户是否为VIP用户"),
        # 故意造错：i(1) 不允许用于代码枚举类
        make_row(table="客户信息明细表", field="支付方式", domain="i(1)",
                 example="1", ftype="代码枚举类", is_enum="是",
                 enum_vals="1.现金 2.银行卡 3.支票 4.汇款",
                 meaning="记录客户选择的支付方式类型"),
        # 故意造错：数据示例"男"含汉字，不符合 an..(1)
        make_row(table="客户信息明细表", field="性别", domain="an..(1)",
                 example="男", ftype="代码枚举类", is_enum="是",
                 enum_vals="男/女", meaning="标识客户的性别信息"),
        # 应通过：见名知意 + 业务含义有实质描述
        make_row(table="客户信息明细表", field="信用等级", domain="an..(1)",
                 example="A", ftype="代码枚举类", is_enum="是",
                 enum_vals="(A)优秀 (B)良好 (C)一般 (D)较差",
                 meaning="反映客户的信用评级，由风控系统定期计算更新"),
        # 故意造错：数据示例"编码"含汉字，不符合 an..(10) + 是否枚举应为否
        make_row(table="客户信息明细表", field="币种代码", domain="an..(10)",
                 example="编码", ftype="编码类", is_enum="是",
                 enum_vals="01：人民币 02：美元 03：欧元",
                 meaning="标识交易使用的货币类型代码，遵循ISO 4217标准"),
        # 故意造错：a..(2) 不允许用于标志类 + 是否枚举应为否但填了枚举值
        make_row(table="客户信息明细表", field="是否冻结", domain="a..(2)",
                 example="Y", ftype="标志类", is_enum="否",
                 enum_vals="Y-是;N-否",
                 meaning="标识账户是否被冻结，冻结后无法进行交易操作"),
        # 故意造错：域类型缺失
        make_row(table="客户信息明细表", field="交易状态", example="01",
                 ftype="代码枚举类", is_enum="是",
                 enum_vals="01-成功;02-失败;03-处理中",
                 meaning="记录交易的当前处理状态"),

        # ==================== 新增应通过样例（10条） ====================

        # --- 客户信息表 ---
        make_row(table="客户信息表", field="客户编号", domain="n..(20)",
                 example="100001", ftype="编码类", is_enum="否",
                 meaning="客户在系统中的唯一标识编号，用于关联客户各类业务信息"),
        make_row(table="客户信息表", field="证件号码", domain="an..(18)",
                 example="110101199001011234", ftype="文本类", is_enum="否",
                 meaning="客户有效身份证件号码，用于实名认证与身份核验"),
        make_row(table="客户信息表", field="联系地址", domain="anc..(200)",
                 example="北京市朝阳区建国路88号", ftype="文本类", is_enum="否",
                 meaning="客户常住地址，用于寄送对账单与通知"),
        make_row(table="客户信息表", field="客户性别", domain="n!(1)",
                 example="1", ftype="代码枚举类", is_enum="是",
                 enum_vals="1-男;2-女",
                 meaning="标识客户性别，1-男，2-女"),

        # --- 交易流水表（新表） ---
        make_row(table="交易流水表", field="交易流水号", domain="an..(32)",
                 example="TX20240115000001", ftype="编码类", is_enum="否",
                 meaning="交易流水的唯一标识编号，由系统自动生成"),
        make_row(table="交易流水表", field="交易金额", domain="i(19,2)",
                 example="1000.50", ftype="数值类", is_enum="否",
                 meaning="交易实际发生金额，单位为元，保留两位小数"),
        make_row(table="交易流水表", field="交易日期", domain="DATE",
                 example="20240115", ftype="日期时间类", is_enum="否",
                 meaning="交易发生的自然日，格式为YYYYMMDD"),

        # --- 贷款信息表（新表） ---
        make_row(table="贷款信息表", field="贷款合同号", domain="an..(32)",
                 example="HT2024010001", ftype="编码类", is_enum="否",
                 meaning="贷款合同的唯一标识编号，关联授信审批与放款记录"),
        make_row(table="贷款信息表", field="贷款金额", domain="i(19,2)",
                 example="500000.00", ftype="数值类", is_enum="否",
                 meaning="贷款合同约定的本金金额，单位为元"),
        make_row(table="贷款信息表", field="贷款状态", domain="n!(1)",
                 example="1", ftype="代码枚举类", is_enum="是",
                 enum_vals="1-正常;2-逾期;3-结清",
                 meaning="贷款当前所处状态，1-正常，2-逾期，3-结清"),
    ]

    df = pd.DataFrame(data)

    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    output = os.path.join(project_root, "data", "qc_input.xlsx")
    df.to_excel(output, index=False)

    # 第1行插入合并表头："基本信息"（跨26列）+ "数管反馈"（跨3列）
    wb = load_workbook(output)
    ws = wb.active
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=26)
    ws.cell(row=1, column=1, value="基本信息")
    ws.merge_cells(start_row=1, start_column=27, end_row=1, end_column=29)
    ws.cell(row=1, column=27, value="数管反馈")
    wb.save(output)

    print(f"示例文件已生成: {output}（{len(data)} 行测试数据）")


if __name__ == "__main__":
    main()
