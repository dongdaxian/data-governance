# -*- coding: utf-8 -*-

"""Excel 读取/写入工具。

读取：将 Excel 解析为 RowData 列表，按精确列名匹配。

  输入 Excel 格式：第1行合并表头"基本信息"，第2行列名，第3行+数据。

写入：在原始数据基础上填写检查结果，输出新 Excel。

  枚举值规范化结果直接覆盖"枚举值(选填)"列；另填"检查结果"、"说明"两列。

"""

import pandas as pd


from quality_check.constants import (

    INPUT_COLUMNS,

    COL_CHECK_RESULT,

    COL_FAIL_REASON,

)

from quality_check.state import RowData

def _match_columns(df: pd.DataFrame) -> dict[str, str]:
    """精确匹配 Excel 列名到内部字段名。

    申请单格式固定，按 INPUT_COLUMNS 中定义的精确列名直接查找。

    Returns:
        {"table_name": "实际列名", "field_name": "...", ...}
    """
    actual_cols = set(str(c) for c in df.columns)
    col_map = {}
    for internal_name, keywords in INPUT_COLUMNS.items():
        for kw in keywords:
            if kw in actual_cols:
                col_map[internal_name] = kw
                break
    return col_map

def _safe_str(val) -> str:

    """安全转字符串，处理 NaN 和空值。"""

    if val is None:

        return ""

    s = str(val).strip()

    if s.lower() in ("nan", "none", "nat"):

        return ""

    return s

def read_excel(file_path: str) -> list[RowData]:

    """读取 Excel 文件，返回 RowData 列表。

    输入 Excel 第1行为合并表头，第2行为列名，从第3行开始为数据。

    """

    df = pd.read_excel(file_path, header=1)

    col_map = _match_columns(df)

    # 校验必填列是否全部匹配到
    required_keys = ["table_name", "field_name", "field_type",
                     "domain_type", "data_example", "is_enum",
                     "business_meaning"]
    missing = [k for k in required_keys if k not in col_map]
    if missing:
        expected = {
            "table_name": "中文表名(必填)",
            "field_name": "中文字段名(必填)",
            "field_type": "字段所属类型(必填)",
            "domain_type": "域类型(必填)",
            "data_example": "数据示例(必填)",
            "is_enum": "是否枚举(必填)",
            "business_meaning": "业务定义(必填)",
        }
        missing_names = [expected[k] for k in missing]
        raise ValueError(
            f"输入Excel缺少必填列: {missing_names}。"
            f"请检查列名是否与申请单模板一致。"
        )

    rows: list[RowData] = []

    for idx, row in df.iterrows():

        rows.append(RowData(

            index=int(idx),

            table_name=_safe_str(row.get(col_map.get("table_name", ""), "")),

            field_name=_safe_str(row.get(col_map.get("field_name", ""), "")),

            field_type=_safe_str(row.get(col_map.get("field_type", ""), "")),

            domain_type=_safe_str(row.get(col_map.get("domain_type", ""), "")),

            data_example=_safe_str(row.get(col_map.get("data_example", ""), "")),

            is_enum=_safe_str(row.get(col_map.get("is_enum", ""), "")),

            business_meaning=_safe_str(row.get(col_map.get("business_meaning", ""), "")),

            enum_values=_safe_str(row.get(col_map.get("enum_values", ""), "")),

            # 初始化结果字段

            rule_issues=[],

            rule_passed=True,

            is_meaningful=True,

            meaning_reason="",

            normalized_enum="",

            flag_issues=[],

            check_result="",

            fail_reason="",

        ))

    print(f"已读取 {len(rows)} 行数据")

    print(f"  列映射: {col_map}")

    return rows

def write_excel(file_path: str, rows: list[RowData], input_file: str):

    """将检查结果写入 Excel。

    完整复制输入文件作为输出（保留全部原始列、数据、合并表头和格式），

    按第2行列名定位"枚举值(选填)/检查结果/说明"三列，从第3行起填入：

    枚举值列直接写入规范化后的枚举值（无规范化结果时保留原值）。

    """

    import shutil

    from openpyxl import load_workbook

    enum_col = INPUT_COLUMNS["enum_values"][0]

    # 1. 完整复制输入文件作为输出文件
    shutil.copyfile(input_file, file_path)

    # 2. 打开输出文件，从第2行（列名行）按列名精确匹配定位结果列
    wb = load_workbook(file_path)
    ws = wb.active

    col_index: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        name = _safe_str(ws.cell(row=2, column=col).value)
        if name in (enum_col, COL_CHECK_RESULT, COL_FAIL_REASON):
            col_index[name] = col

    missing = [n for n in (enum_col, COL_CHECK_RESULT, COL_FAIL_REASON)
               if n not in col_index]
    if missing:
        raise ValueError(
            f"输入Excel缺少必填列: {missing}，请检查模板是否与申请单格式一致。"
        )

    # 3. 从第3行起按 index 升序逐行填入结果
    sorted_rows = sorted(rows, key=lambda r: r["index"])
    for r in sorted_rows:
        excel_row = 3 + r["index"]
        ws.cell(
            row=excel_row, column=col_index[enum_col],
            value=r["normalized_enum"] or r["enum_values"],
        )
        ws.cell(row=excel_row, column=col_index[COL_CHECK_RESULT], value=r["check_result"])
        ws.cell(row=excel_row, column=col_index[COL_FAIL_REASON], value=r["fail_reason"])

    wb.save(file_path)

    print(f"结果已写入: {file_path}")

    # 打印统计

    total = len(sorted_rows)

    passed = sum(1 for r in sorted_rows if r["check_result"] == "通过")

    failed = total - passed

    print(f"  总计: {total} 行 | 通过: {passed} | 不通过: {failed}")

