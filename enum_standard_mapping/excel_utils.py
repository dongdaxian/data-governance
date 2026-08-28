# -*- coding: utf-8 -*-
"""Excel 读取/写入工具（枚举落标）。

读取：解析质检输出格式 Excel，筛选代码枚举类字段。
写入：完整复制输入文件，追加七分类结果列。
"""

import shutil

import pandas as pd
from openpyxl import load_workbook

from enum_standard_mapping.constants import (
    INPUT_COLUMNS,
    ENUM_FIELD_TYPE,
    COL_MAPPING_RESULT,
    COL_SELECTED_STD_ID,
    COL_SELECTED_STD_NAME,
    COL_DOMAIN_ACTION,
    COL_CONFLICT_DETAIL,
    COL_LLM_REASON,
    COL_CANDIDATES,
)
from enum_standard_mapping.state import FieldToMap


def _safe_str(val) -> str:
    """安全转字符串，处理 NaN 和空值。"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    return s


def _match_columns(df: pd.DataFrame) -> dict[str, str]:
    """精确匹配 Excel 列名到内部字段名。"""
    actual_cols = set(str(c) for c in df.columns)
    col_map = {}
    for internal_name, keywords in INPUT_COLUMNS.items():
        for kw in keywords:
            if kw in actual_cols:
                col_map[internal_name] = kw
                break
    return col_map


def read_excel(file_path: str) -> list[FieldToMap]:
    """读取 Excel（质检输出格式），返回全部行（未筛选，筛选在节点内做）。"""
    df = pd.read_excel(file_path, header=1)
    col_map = _match_columns(df)

    required_keys = ["field_name", "field_type", "business_meaning", "enum_values"]
    missing = [k for k in required_keys if k not in col_map]
    if missing:
        raise ValueError(
            f"输入Excel缺少必填列: {missing}，"
            f"请检查输入文件是否为质检输出格式（需含'格式化枚举值'列）。"
        )

    rows: list[FieldToMap] = []
    for idx, row in df.iterrows():
        rows.append(FieldToMap(
            index=int(idx),
            field_name=_safe_str(row.get(col_map["field_name"], "")),
            field_type=_safe_str(row.get(col_map["field_type"], "")),
            business_meaning=_safe_str(row.get(col_map["business_meaning"], "")),
            enum_values=_safe_str(row.get(col_map["enum_values"], "")),
            # 初始化结果字段
            pairs=[],
            candidates=[],
            degraded_name_hits=[],
            candidate_fetch_error="",
            mapping_result="",
            selected_std_id="",
            selected_std_name="",
            domain_action="",
            conflict_detail="",
            llm_reason="",
        ))

    print(f"已读取 {len(rows)} 行数据")
    print(f"  列映射: {col_map}")
    return rows


def _format_candidates(candidates) -> str:
    """将候选枚举值项格式化为明细文本（测试输出用）。"""
    if not candidates:
        return ""
    parts = []
    for c in candidates:
        parts.append(
            f"{c['item_id']} {c['item_name']}({c['item_type']},得分{c['score']}/n,"
            f"字典{'+'.join(s['std_id'] for s in c['standards'])})"
        )
    return "; ".join(parts)


def write_excel(
    file_path: str,
    rows: list[FieldToMap],
    input_file: str,
    include_candidates: bool = False,
):
    """将枚举落标结果写入 Excel。

    完整复制输入文件作为输出，末尾追加结果列。
    非代码枚举类字段（未处理行）结果列填空。
    """
    shutil.copyfile(input_file, file_path)

    wb = load_workbook(file_path)
    ws = wb.active
    max_col = ws.max_column

    result_cols = [
        COL_MAPPING_RESULT,
        COL_SELECTED_STD_ID,
        COL_SELECTED_STD_NAME,
        COL_DOMAIN_ACTION,
        COL_CONFLICT_DETAIL,
        COL_LLM_REASON,
    ]
    if include_candidates:
        result_cols.append(COL_CANDIDATES)

    # 第2行追加列名，第1行追加合并表头
    col_offset = max_col + 1
    for i, col_name in enumerate(result_cols):
        ws.cell(row=2, column=col_offset + i, value=col_name)

    start_col = col_offset
    end_col = col_offset + len(result_cols) - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    ws.cell(row=1, column=start_col, value="落标结果")

    # 从第3行起逐行填入
    df = pd.read_excel(input_file, header=1)
    n = len(df)
    row_map = {r["index"]: r for r in rows}

    for idx in range(n):
        excel_row = 3 + idx
        r = row_map.get(idx)
        if r is not None:
            ws.cell(row=excel_row, column=col_offset, value=r["mapping_result"])
            ws.cell(row=excel_row, column=col_offset + 1, value=r["selected_std_id"])
            ws.cell(row=excel_row, column=col_offset + 2, value=r["selected_std_name"])
            ws.cell(row=excel_row, column=col_offset + 3, value=r["domain_action"])
            ws.cell(row=excel_row, column=col_offset + 4, value=r["conflict_detail"])
            ws.cell(row=excel_row, column=col_offset + 5, value=r["llm_reason"])
            if include_candidates:
                ws.cell(row=excel_row, column=col_offset + 6, value=_format_candidates(r["candidates"]))
        else:
            for i in range(len(result_cols)):
                ws.cell(row=excel_row, column=col_offset + i, value="")

    wb.save(file_path)
    print(f"结果已写入: {file_path}")

    # 打印统计
    from enum_standard_mapping.constants import (
        RESULT_1, RESULT_2, RESULT_3, RESULT_4,
        RESULT_5, RESULT_6, RESULT_7,
    )
    sorted_rows = sorted(rows, key=lambda r: r["index"])
    total = len(sorted_rows)
    stats = {
        RESULT_1: 0, RESULT_2: 0, RESULT_3: 0, RESULT_4: 0,
        RESULT_5: 0, RESULT_6: 0, RESULT_7: 0,
    }
    other = 0
    for r in sorted_rows:
        if r["mapping_result"] in stats:
            stats[r["mapping_result"]] += 1
        else:
            other += 1
    print(f"  总计: {total} 行 | " + " | ".join(f"{k}: {v}" for k, v in stats.items()) + f" | 其他: {other}")
