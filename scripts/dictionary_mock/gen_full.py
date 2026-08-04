# -*- coding: utf-8 -*-
"""好东西全量字典处理: 域匹配 + 代码枚举类枚举值生成 + 质检 + 导出
用法:
  python gen_full.py --phase nonenum    # 非枚举域匹配(规则+LLM)
  python gen_full.py --phase codeenum   # 代码枚举类处理(名修复+枚举值生成)
  python gen_full.py --phase export     # 质检+导出
  python gen_full.py --phase all        # 全流程
"""
import sys, os, json, re, time, random, argparse, threading
sys.stdout.reconfigure(encoding="utf-8")
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
DICT_PATH = os.path.join(HERE, "全量字典.xlsx")
DOM_PATH = os.path.join(HERE, "域表.xlsx")
CACHE_PATH = os.path.join(HERE, "gen_cache.json")

API_KEYS = os.environ.get("API_KEY", "").split(",")
if not API_KEYS or not API_KEYS[0]:
    API_KEYS = ["REPLACE_WITH_YOUR_KEY"]
_key_idx = 0
_key_lock = threading.Lock()
URL = "https://ark.cn-beijing.volces.com/api/coding/v3/chat/completions"
MODEL = "GLM-5.2"
WORKERS = 6

SHEET_TYPES = ["日期时间类","标志类","编码类","数值类","文本类","代码枚举类"]
OUT_COLS = ["标准编号","标准中文名称","标准英文全称","标准物理字段名","标准物理字段类型",
            "标准所属类型","业务定义","业务口径","业务规则",
            "域名称","域编号","域类型","枚举值编号"]

# ============ LLM ============
def _cur_key():
    with _key_lock:
        return API_KEYS[_key_idx % len(API_KEYS)]
def _next_key():
    global _key_idx
    with _key_lock:
        _key_idx += 1
        return API_KEYS[_key_idx % len(API_KEYS)]

def call_llm(prompt, max_tokens=6000, timeout=150):
    payload = {"model": MODEL, "messages":[{"role":"user","content":prompt}],
               "max_tokens": max_tokens, "temperature": 0.3, "thinking":{"type":"disabled"}}
    last_err = None
    for attempt in range(4):
        key = _cur_key()
        headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
        r = None
        try:
            r = requests.post(URL, json=payload, headers=headers, timeout=timeout)
            if r.status_code in (401, 403, 429):
                print(f"  key#{_key_idx%len(API_KEYS)} HTTP{r.status_code}, 切换", flush=True)
                _next_key()
                last_err = f"HTTP {r.status_code}"
                time.sleep(2)
                continue
            r.raise_for_status()
            return r.json()["choices"][0]["message"].get("content","") or ""
        except Exception as e:
            last_err = f"{type(e).__name__} {str(e)[:80]}"
            if r is not None and r.status_code in (401, 403, 429):
                _next_key(); time.sleep(2); continue
            if attempt < 3: time.sleep(3 * (attempt + 1))
    print(f"  LLM失败: {last_err}", flush=True)
    return ""

def extract_json(text):
    text = (text or "").strip()
    text = re.sub(r"^```(json)?", "", text).strip()
    text = re.sub(r"```$", "", text).strip()
    m = re.search(r"\[.*\]", text, re.S)
    if m: text = m.group(0)
    return json.loads(text)

def load_cache():
    if os.path.exists(CACHE_PATH):
        return json.load(open(CACHE_PATH, encoding="utf-8"))
    return {}
def save_cache(cache):
    json.dump(cache, open(CACHE_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

STATE_PATH = os.path.join(HERE, "gen_dict_state.json")
def save_state(dictionary, new_codeenum_domains=None):
    state = {"dictionary": dictionary, "new_codeenum_domains": new_codeenum_domains or []}
    json.dump(state, open(STATE_PATH, "w", encoding="utf-8"), ensure_ascii=False)
    print(f"  状态已保存: {STATE_PATH}", flush=True)
def load_state():
    if os.path.exists(STATE_PATH):
        state = json.load(open(STATE_PATH, encoding="utf-8"))
        print(f"  载入已有状态", flush=True)
        return state["dictionary"], state.get("new_codeenum_domains", [])
    return None, []

# ============ 数据载入 ============
def load_dictionary():
    """返回 {sheet_type: [entry_dict, ...]}"""
    wb = load_workbook(DICT_PATH, read_only=True)
    result = {}
    for ws in wb.worksheets:
        if ws.title not in SHEET_TYPES:
            continue
        headers = [c.value for c in ws[1]]
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for i, h in enumerate(headers):
                if h:
                    d[h] = row[i] if i < len(row) else None
            d["_sheet"] = ws.title
            entries.append(d)
        result[ws.title] = entries
        print(f"  载入 {ws.title}: {len(entries)} 条", flush=True)
    wb.close()
    return result

def load_domains():
    """返回 {域编号: {域编号,域名称,域类型,标准所属类型,枚举值编号}}"""
    wb = load_workbook(DOM_PATH, read_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    domains = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        did = d.get("域编号")
        if did:
            domains[did] = d
    wb.close()
    # Build name->domain lookup by type
    by_type = {}
    for d in domains.values():
        cat = d.get("标准所属类型","")
        by_type.setdefault(cat, []).append(d)
    return domains, by_type

# ============ 非枚举域匹配 ============
def match_nonenum_rulebased(dictionary, domains_by_type):
    """规则匹配: 日期时间类/标志类/编码类已有域名的"""
    # Build name->domain lookup per type
    name_lookup = {}
    for cat, doms in domains_by_type.items():
        name_lookup[cat] = {d["域名称"]: d for d in doms}

    matched = 0
    unmatched = 0

    for sheet_type in ["日期时间类","标志类","编码类"]:
        entries = dictionary.get(sheet_type, [])
        nl = name_lookup.get(sheet_type, {})
        for e in entries:
            dom_name = e.get("域", "")
            if dom_name and dom_name in nl:
                d = nl[dom_name]
                e["域名称"] = d["域名称"]
                e["域编号"] = d["域编号"]
                e["域类型"] = d["域类型"]
                e["枚举值编号"] = d.get("枚举值编号","")
                matched += 1
            elif dom_name == "机构号":
                # 机构号 -> 通用编号
                d = nl.get("通用编号", {}) or name_lookup.get("编码类",{}).get("通用编号",{})
                if d:
                    e["域名称"] = "通用编号"
                    e["域编号"] = d["域编号"]
                    e["域类型"] = d["域类型"]
                    e["枚举值编号"] = ""
                    matched += 1
                else:
                    unmatched += 1
            elif dom_name == "时间戳":
                # 时间戳 -> 随机选 时间戳1 或 时间戳2
                ts_doms = [d for d in domains_by_type.get("日期时间类",[]) if "时间戳" in d.get("域名称","")]
                if ts_doms:
                    d = random.choice(ts_doms)
                    e["域名称"] = d["域名称"]
                    e["域编号"] = d["域编号"]
                    e["域类型"] = d["域类型"]
                    e["枚举值编号"] = ""
                    matched += 1
                else:
                    unmatched += 1
            elif dom_name:
                # 有域名但不在域表中，尝试模糊匹配
                unmatched += 1
            else:
                unmatched += 1
        print(f"  {sheet_type}: 规则匹配 {sum(1 for e in entries if e.get('域编号'))}/{len(entries)}", flush=True)

    # 标志类全部用 标志
    for e in dictionary.get("标志类", []):
        if not e.get("域编号"):
            flg = name_lookup.get("标志类",{}).get("标志",{})
            if flg:
                e["域名称"] = "标志"
                e["域编号"] = flg["域编号"]
                e["域类型"] = flg["域类型"]
                e["枚举值编号"] = ""

    return matched, unmatched

def build_match_prompt(entries, doms, cat):
    dom_lines = "; ".join(f"{d['域编号']}={d['域名称']}({d['域类型']})" for d in doms)
    items = []
    for i, e in enumerate(entries):
        cn = e.get("标准中文名称","") or ""
        definition = (e.get("业务定义","") or "")[:80]
        items.append(f'{i}. {cn} - {definition}')
    joined = "\n".join(items)
    return f"""你是银行业数据标准专家。请为以下【{cat}】字段匹配最合适的域。

【可选域清单】（只能选以下域编号）
{dom_lines}

【匹配规则】
1. 根据字段名称和业务含义选择最匹配的域
2. 如果字段叫"**编号"且有对应的域(如"任务编号"->任务编号域)，优先用对应域
3. 如果没有完全对应的，用"通用编号"(编码类)或最通用的域
4. 如果多个域都适合(如名称1/名称2/名称3)，随机选一个
5. 金额类字段选金额相关域，余额类选余额域，利率类选利率域

【字段列表】
{joined}

只输出JSON数组，不要多余文字：
[{{"idx":0,"域编号":"ECD00005"}},{{"idx":1,"域编号":"NUM00001"}}]
"""

def match_nonenum_llm(dictionary, domains_by_type, cache):
    """LLM匹配: 编码类未匹配 + 数值类 + 文本类"""
    rng = random.Random(42)
    for cat in ["编码类","数值类","文本类"]:
        entries = dictionary.get(cat, [])
        doms = domains_by_type.get(cat, [])
        if not doms:
            print(f"  {cat}: 无可用域，跳过", flush=True)
            continue

        # 收集需要LLM匹配的条目
        todo = [(i, e) for i, e in enumerate(entries) if not e.get("域编号")]
        if not todo:
            print(f"  {cat}: 全部已匹配", flush=True)
            continue

        cache_key = f"nonenum_{cat}"
        cached = cache.get(cache_key, {})
        todo_ids = [idx for idx, _ in todo if str(idx) not in cached]

        print(f"  {cat}: 待匹配 {len(todo)} 条 (缓存 {len(cached)}, 新 {len(todo_ids)})", flush=True)

        # 分批
        batches = []
        for j in range(0, len(todo_ids), 30):
            batch_ids = todo_ids[j:j+30]
            batch_entries = [entries[idx] for idx in batch_ids]
            batches.append((batch_ids, batch_entries))

        def _do(batch):
            batch_ids, batch_entries = batch
            try:
                text = call_llm(build_match_prompt(batch_entries, doms, cat), max_tokens=2000)
                arr = extract_json(text)
                local = {}
                for item in arr:
                    idx = item.get("idx")
                    if idx is not None and 0 <= idx < len(batch_ids):
                        did = item.get("域编号","")
                        local[str(batch_ids[idx])] = did
                return local
            except Exception as e:
                print(f"  {cat} 批次异常: {str(e)[:80]}", flush=True)
                return {}

        done = 0
        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            futs = {ex.submit(_do, b): b for b in batches}
            for fut in as_completed(futs):
                done += 1
                local = fut.result()
                cached.update(local)
                cache[cache_key] = cached
                if done % 10 == 0 or done == len(batches):
                    save_cache(cache)
                    print(f"  {cat} 进度 {done}/{len(batches)}", flush=True)
        save_cache(cache)

        # 应用结果
        matched_cnt = 0
        for idx, e in enumerate(entries):
            did = cached.get(str(idx))
            if did and did in {d["域编号"] for d in doms}:
                d = next(d for d in doms if d["域编号"] == did)
                e["域名称"] = d["域名称"]
                e["域编号"] = did
                e["域类型"] = d["域类型"]
                e["枚举值编号"] = ""
                matched_cnt += 1
        print(f"  {cat}: LLM匹配 {matched_cnt}/{len(entries)}", flush=True)

# ============ 代码枚举类处理 ============
def fix_codeenum_names(entries):
    """给不以"代码"结尾的名称加上"代码"后缀，同步修改英文名和物理名"""
    fixed = 0
    for e in entries:
        cn = e.get("标准中文名称","") or ""
        if not cn.endswith("代码"):
            e["标准中文名称"] = cn + "代码"
            # 英文名加 CODE
            en = e.get("标准英文全称","") or ""
            if not en.upper().rstrip().endswith("CODE"):
                e["标准英文全称"] = (en.rstrip() + " CODE").strip()
            # 物理名加 _CD
            ph = e.get("标准物理字段名","") or ""
            if not ph.upper().endswith("CD"):
                new_ph = ph.rstrip("_") + "_CD"
                e["标准物理字段名"] = new_ph[:24]
            fixed += 1
    print(f"  代码枚举类名修复: {fixed}/{len(entries)} 条加了'代码'后缀", flush=True)
    return fixed

def format_enum_rule(rule_text):
    """尝试规则化整理已有的业务规则为 1-x;2-y 格式"""
    if not rule_text:
        return None
    rule = str(rule_text).strip()
    # 检测是否是行标国标引用
    if re.search(r'参考|参照|参见|《.*》|GB/?T|JR/T|行标|国标', rule):
        return rule  # 保留原样
    # 尝试提取 数字-名称 模式
    # 匹配 "01-正常" "1--禁止" "01、正常" 等
    patterns = re.findall(r'(\d+)\s*[-—、,，:：]\s*([^\d;,，\n；]+)', rule)
    if len(patterns) >= 2:
        parts = [f"{code}-{name.strip()}" for code, name in patterns]
        return ";".join(parts)
    return None  # 无法规则化，需要LLM

def build_enum_prompt(entries):
    items = []
    for i, e in enumerate(entries):
        cn = e.get("标准中文名称","") or ""
        definition = (e.get("业务定义","") or "")[:100]
        items.append(f'{i}. {cn} - {definition}')
    joined = "\n".join(items)
    return f"""你是银行业数据标准专家。请为以下代码枚举类字段生成枚举值。

【字段列表】
{joined}

【要求】
1. 根据字段名称和业务含义生成合理的枚举值
2. 格式为"编码-名称;编码-名称"（如"1-男;2-女"或"01-正常;02-冻结;03-注销"）
3. 枚举值数量2-10个，根据实际业务场景确定
4. 编码简洁（数字或字母），名称简短
5. 只输出业务规则字符串，不要多余解释

只输出JSON数组：[{{"idx":0,"业务规则":"1-男;2-女"}},{{"idx":1,"业务规则":"01-正常;02-冻结;03-注销"}}]
"""

def process_codeenum(dictionary, cache):
    entries = dictionary.get("代码枚举类", [])
    print(f"  代码枚举类: {len(entries)} 条", flush=True)

    # 1. 修复名称
    fix_codeenum_names(entries)

    # 2. 处理业务规则
    need_llm = []  # (idx, entry)
    for i, e in enumerate(entries):
        rule = e.get("业务规则")
        formatted = format_enum_rule(rule)
        if formatted is not None:
            e["业务规则"] = formatted
        else:
            need_llm.append((i, e))

    print(f"  业务规则: 规则化/保留 {len(entries)-len(need_llm)}, 需LLM {len(need_llm)}", flush=True)

    # 3. LLM生成
    cache_key = "codeenum_rules"
    cached = cache.get(cache_key, {})
    todo = [(i, e) for i, e in need_llm if str(i) not in cached]
    print(f"  缓存 {len(cached)}, 新 {len(todo)}", flush=True)

    batches = []
    for j in range(0, len(todo), 15):
        batch = todo[j:j+15]
        batches.append(batch)

    def _do(batch):
        batch_entries = [e for _, e in batch]
        try:
            text = call_llm(build_enum_prompt(batch_entries), max_tokens=3000)
            arr = extract_json(text)
            local = {}
            for item in arr:
                idx = item.get("idx")
                if idx is not None and 0 <= idx < len(batch):
                    rule = item.get("业务规则","")
                    local[str(batch[idx][0])] = rule
            return local
        except Exception as e:
            print(f"  枚举值批次异常: {str(e)[:80]}", flush=True)
            return {}

    done = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = {ex.submit(_do, b): b for b in batches}
        for fut in as_completed(futs):
            done += 1
            local = fut.result()
            cached.update(local)
            cache[cache_key] = cached
            if done % 20 == 0 or done == len(batches):
                save_cache(cache)
                print(f"  枚举值进度 {done}/{len(batches)}", flush=True)
    save_cache(cache)

    # 应用结果
    for i, e in enumerate(entries):
        rule = cached.get(str(i))
        if rule:
            e["业务规则"] = rule

    # 4. 创建域和枚举值编号
    # 域名 = 中文名去掉"代码"后缀
    cde_counter = 1
    enm_counter = 1
    new_domains = []
    for e in entries:
        cn = e.get("标准中文名称","")
        dom_name = cn[:-2] if cn.endswith("代码") else cn
        # 确定域类型(根据枚举值编码长度)
        rule = e.get("业务规则","") or ""
        codes = re.findall(r'^(\S+?)-', rule + ";", re.M)
        # 找所有 "xxx-" 的 xxx 部分 (编码部分)
        code_matches = re.findall(r'(?:^|;)([^\s;-]+)-', rule)
        max_len = max((len(c) for c in code_matches), default=2)
        dtype = f"an..({max(max_len,2)})"
        if "参考" in rule or "《" in rule:
            dtype = "an..(20)"

        cde_id = f"CDE{cde_counter:05d}"
        enm_id = f"ENM{enm_counter:05d}"
        e["域名称"] = dom_name
        e["域编号"] = cde_id
        e["域类型"] = dtype
        e["枚举值编号"] = enm_id
        new_domains.append({
            "域编号": cde_id, "域名称": dom_name, "域类型": dtype,
            "标准所属类型": "代码枚举类", "枚举值编号": enm_id
        })
        cde_counter += 1
        enm_counter += 1

    print(f"  创建代码枚举域: {len(new_domains)} 个", flush=True)
    return new_domains

# ============ 质检 ============
def quality_check(dictionary):
    """检查类型/名称一致性，删除不匹配项"""
    deleted = 0
    for sheet_type, entries in dictionary.items():
        before = len(entries)
        kept = []
        for e in entries:
            cn = e.get("标准中文名称","") or ""
            cat = e.get("标准所属类型","") or ""
            # 检查: 标志类名称不应出现在其他类型中(反向也检查)
            if cat != "标志类" and cn.endswith("标志") and not cn.endswith("代码"):
                continue  # 删除: 非标志类但叫"**标志"
            if cat == "标志类" and not cn.endswith("标志"):
                # 标志类但不叫标志? 保留(可能是特殊情况)
                pass
            # 检查: 无域编号的删除(除非是代码枚举类已处理)
            if not e.get("域编号") and cat != "代码枚举类":
                continue  # 删除: 无域匹配
            kept.append(e)
        deleted += before - len(kept)
        dictionary[sheet_type] = kept
        if before != len(kept):
            print(f"  {sheet_type}: {before} -> {len(kept)} (删 {before-len(kept)})", flush=True)
    print(f"  质检删除: {deleted} 条", flush=True)
    return deleted

# ============ 导出 ============
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="4472C4")
WRAP = Alignment(wrap_text=True, vertical="top")

def style_sheet(ws, widths=None):
    for c in ws[1]:
        c.font = HDR_FONT; c.fill = HDR_FILL
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

def export_all(dictionary, original_domains, new_codeenum_domains):
    # 1. 全量字典_完整.xlsx (6 sheets + 新列)
    out_dict = os.path.join(HERE, "全量字典_完整.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_type in SHEET_TYPES:
        entries = dictionary.get(sheet_type, [])
        ws = wb.create_sheet(sheet_type)
        ws.append(OUT_COLS)
        for e in entries:
            ws.append([e.get(k,"") for k in OUT_COLS])
        style_sheet(ws, {"A":12,"B":22,"C":36,"D":22,"E":18,"F":10,"G":40,"H":20,"I":20,"J":20,"K":12,"L":12,"M":12})
    wb.save(out_dict)
    print(f"导出: {out_dict} ({sum(len(v) for v in dictionary.values())} 条)", flush=True)

    # 2. 域表_完整.xlsx (89原 + 新增代码枚举域)
    out_dom = os.path.join(HERE, "域表_完整.xlsx")
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet1"
    h2 = ["域编号","域名称","域类型","标准所属类型","枚举值编号"]
    ws2.append(h2)
    for d in original_domains.values():
        ws2.append([d.get(k,"") for k in h2])
    for d in new_codeenum_domains:
        ws2.append([d.get(k,"") for k in h2])
    style_sheet(ws2, {"A":12,"B":30,"C":15,"D":12,"E":12})
    wb2.save(out_dom)
    print(f"导出: {out_dom} ({len(original_domains)+len(new_codeenum_domains)} 域)", flush=True)

    # 3. 枚举值表 + 枚举值编码明细
    out_enum = os.path.join(HERE, "枚举值表.xlsx")
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "枚举值表"
    h3 = ["枚举值编号","枚举值名称","编码规则","域编号","域类型","枚举值数量"]
    ws3.append(h3)
    ws4 = wb3.create_sheet("枚举值编码明细")
    h4 = ["枚举值编号","枚举编码","枚举名称"]
    ws4.append(h4)

    for e in dictionary.get("代码枚举类", []):
        enm = e.get("枚举值编号","")
        dom_name = e.get("域名称","")
        cde = e.get("域编号","")
        dtype = e.get("域类型","")
        rule = e.get("业务规则","") or ""
        # 解析枚举值
        codes = []
        if "参考" in rule or "《" in rule:
            # 行标国标，整体作为编码规则
            ws3.append([enm, dom_name, rule, cde, dtype, 0])
        else:
            # 解析 "1-男;2-女" 格式
            parts = rule.split(";") if ";" in rule else rule.split("，")
            for p in parts:
                p = p.strip()
                if "-" in p:
                    code, name = p.split("-", 1)
                    code = code.strip()
                    name = name.strip()
                    if code and name:
                        codes.append((code, name))
            ws3.append([enm, dom_name, rule, cde, dtype, len(codes)])
            for code, name in codes:
                ws4.append([enm, code, name])

    style_sheet(ws3, {"A":12,"B":30,"C":50,"D":12,"E":12,"F":10})
    style_sheet(ws4, {"A":12,"B":12,"C":20})
    wb3.save(out_enum)
    print(f"导出: {out_enum}", flush=True)

# ============ 主流程 ============
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--phase", default="all")
    args = ap.parse_args()

    cache = load_cache()
    domains, domains_by_type = load_domains()

    # 尝试载入已有状态(跨阶段续跑)
    saved_dict, saved_domains = load_state()
    if saved_dict and args.phase != "all":
        dictionary = saved_dict
        new_codeenum_domains = saved_domains
        print(f"  从状态恢复: {sum(len(v) for v in dictionary.values())} 条", flush=True)
    else:
        dictionary = load_dictionary()
        new_codeenum_domains = []

    if args.phase in ("all", "nonenum"):
        print("\n=== 非枚举域匹配 ===", flush=True)
        match_nonenum_rulebased(dictionary, domains_by_type)
        match_nonenum_llm(dictionary, domains_by_type, cache)
        save_state(dictionary, new_codeenum_domains)

    if args.phase in ("all", "codeenum"):
        print("\n=== 代码枚举类处理 ===", flush=True)
        new_codeenum_domains = process_codeenum(dictionary, cache)
        save_state(dictionary, new_codeenum_domains)

    if args.phase in ("all", "export"):
        print("\n=== 质检 ===", flush=True)
        quality_check(dictionary)
        print("\n=== 导出 ===", flush=True)
        export_all(dictionary, domains, new_codeenum_domains)

    print("\n完成!", flush=True)

if __name__ == "__main__":
    main()
