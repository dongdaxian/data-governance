# -*- coding: utf-8 -*-
"""最终处理: 代码枚举类枚举值清理(LLM) + 反思检查 + 合并导出4 sheet
用法:
  python gen_final.py --phase all       # 全流程
  python gen_final.py --phase cleanup   # 仅Phase1: LLM清理
  python gen_final.py --phase reflect   # 仅Phase2: 反思检查
  python gen_final.py --phase export    # 仅导出
"""
import sys, os, json, re, time, argparse, threading
sys.stdout.reconfigure(encoding="utf-8")
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
DICT_PATH = os.path.join(HERE, "全量字典_完整.xlsx")
DOM_PATH = os.path.join(HERE, "域表_完整.xlsx")
CACHE_PATH = os.path.join(HERE, "final_cache.json")
STATE_PATH = os.path.join(HERE, "final_state.json")
OUT_PATH = os.path.join(HERE, "全量字典_最终.xlsx")

API_KEYS = os.environ.get("API_KEY", "").split(",")
if not API_KEYS or not API_KEYS[0]:
    API_KEYS = ["REPLACE_WITH_YOUR_KEY"]
_key_idx = 0
_key_lock = threading.Lock()
URL = "https://ark.cn-beijing.volces.com/api/coding/v3/chat/completions"
MODEL = "GLM-5.2"
WORKERS = 6
BATCH = 15

SHEET_TYPES = ["日期时间类","标志类","编码类","数值类","文本类","代码枚举类"]
DICT_COLS = ["标准编号","标准中文名称","标准英文全称","标准物理字段名","标准物理字段类型",
             "标准所属类型","业务定义","业务口径","业务规则",
             "域名称","域编号","域类型","枚举值编号"]

# ============ LLM ============
def _cur_key():
    with _key_lock:
        return API_KEYS[_key_idx % len(API_KEYS)]
def _next_key():
    global _key_idx
    with _key_lock:
        _key_idx += 1
        return API_KEYS[_key_idx % len(API_KEYS)]

def call_llm(prompt, max_tokens=4000, timeout=150):
    payload = {"model": MODEL, "messages":[{"role":"user","content":prompt}],
               "max_tokens": max_tokens, "temperature": 0.3, "thinking":{"type":"disabled"}}
    last_err = None
    for attempt in range(4):
        key = _cur_key()
        headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
        r = None
        try:
            r = requests.post(URL, json=payload, headers=headers, timeout=timeout)
            if r.status_code in (401, 403, 429):
                print(f"  key#{_key_idx%len(API_KEYS)} HTTP{r.status_code}, 切换", flush=True)
                _next_key(); last_err = f"HTTP {r.status_code}"; time.sleep(2); continue
            r.raise_for_status()
            return r.json()["choices"][0]["message"].get("content","") or ""
        except Exception as e:
            last_err = f"{type(e).__name__} {str(e)[:80]}"
            if r is not None and r.status_code in (401, 403, 429):
                _next_key(); time.sleep(2); continue
            if attempt < 3: time.sleep(3 * (attempt + 1))
    print(f"  LLM失败: {last_err}", flush=True)
    return ""

def extract_json(text):
    text = (text or "").strip()
    text = re.sub(r"^```(json)?", "", text).strip()
    text = re.sub(r"```$", "", text).strip()
    m = re.search(r"\[.*\]", text, re.S)
    if m: text = m.group(0)
    return json.loads(text)

def load_cache():
    if os.path.exists(CACHE_PATH):
        return json.load(open(CACHE_PATH, encoding="utf-8"))
    return {}
def save_cache(cache):
    json.dump(cache, open(CACHE_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

def save_state(data):
    json.dump(data, open(STATE_PATH, "w", encoding="utf-8"), ensure_ascii=False)
    print(f"  状态已保存", flush=True)
def load_state():
    if os.path.exists(STATE_PATH):
        return json.load(open(STATE_PATH, encoding="utf-8"))
    return None

# ============ 数据载入 ============
def load_dictionary():
    wb = load_workbook(DICT_PATH, read_only=True)
    result = {}
    for ws in wb.worksheets:
        if ws.title not in SHEET_TYPES: continue
        headers = [c.value for c in ws[1]]
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for i, h in enumerate(headers):
                if h: d[h] = row[i] if i < len(row) else None
            d["_sheet"] = ws.title
            entries.append(d)
        result[ws.title] = entries
        print(f"  载入 {ws.title}: {len(entries)} 条", flush=True)
    wb.close()
    return result

def load_domains():
    wb = load_workbook(DOM_PATH, read_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    domains = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        domains.append(dict(zip(headers, row)))
    wb.close()
    return domains

# ============ Phase 1: LLM 清理 ============
def build_cleanup_prompt(entries):
    items = []
    for i, e in enumerate(entries):
        cn = (e.get("标准中文名称","") or "")[:30]
        df = (e.get("业务定义","") or "")[:80]
        kal = (e.get("业务口径","") or "")[:80]
        rule = (e.get("业务规则","") or "")[:80]
        items.append(f'{i}. [{cn}] 定义:{df} | 口径:{kal} | 规则:{rule}')
    joined = "\n".join(items)
    return f"""你是银行业数据标准专家。请检查以下代码枚举类字段的"业务口径"和"业务规则"列，按规则处理。

【处理规则】
对每个字段，先判断"业务口径"属于哪种情况：

情况1 - 业务口径包含枚举值（如"01--正常;02特殊"或"1-男,2-女"）：
  将业务口径中的枚举值整理为标准格式"编码-名称;编码-名称"（如"01-正常;02-特殊"），覆盖到"业务规则"
  如果业务口径和业务规则的枚举值不同，请根据字段名称和业务含义判断哪组更准确，选择更准确的那组
  "业务口径"清空（输出空字符串）

情况2 - 业务口径为空：
  复查"业务规则"是否与字段名称、业务含义相关
  如果有问题（枚举值不相关或格式错误），重新生成
  如果没问题，保持原样

情况3 - 业务口径是字段来源/口径描述（如"核心贷款主档中的字段值"），不是枚举值：
  "业务口径"保持原样不变
  检查"业务规则"是否与口径一致，格式是否正确
  如果有问题，根据口径和含义重新生成枚举值
  如果没问题，保持原样

【格式要求】
- 业务规则统一为"编码-名称;编码-名称"格式（分号分隔），修复双横线"--"为单横线"-"，去除尾部多余符号
- 如果业务规则引用行标国标（含《...》或"参考"或"GB"），保持原样不改动
- 业务口径在情况1时清空，情况3时保持原样，情况2时保持为空

【字段列表】
{joined}

只输出JSON数组，不要多余文字：
[{{"idx":0,"case":1,"业务规则":"01-正常;02-特殊","业务口径":""}},{{"idx":1,"case":2,"业务规则":"1-男;2-女;0-未知","业务口径":""}}]
"""

def phase_cleanup(dictionary, cache):
    entries = dictionary["代码枚举类"]
    print(f"\n=== Phase 1: LLM清理 ({len(entries)}条) ===", flush=True)
    cache_key = "cleanup"
    cached = cache.get(cache_key, {})
    todo = [(i, e) for i, e in enumerate(entries) if str(i) not in cached]
    print(f"  缓存 {len(cached)}, 新 {len(todo)}", flush=True)

    batches = [todo[j:j+BATCH] for j in range(0, len(todo), BATCH)]

    def _do(batch):
        batch_entries = [e for _, e in batch]
        try:
            text = call_llm(build_cleanup_prompt(batch_entries), max_tokens=4000)
            arr = extract_json(text)
            local = {}
            for item in arr:
                idx = item.get("idx")
                if idx is not None and 0 <= idx < len(batch):
                    local[str(batch[idx][0])] = {
                        "case": item.get("case",2),
                        "业务规则": item.get("业务规则",""),
                        "业务口径": item.get("业务口径",""),
                    }
            return local
        except Exception as e:
            print(f"  清理批次异常: {str(e)[:80]}", flush=True)
            return {}

    done = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = {ex.submit(_do, b): b for b in batches}
        for fut in as_completed(futs):
            done += 1
            local = fut.result()
            cached.update(local)
            cache[cache_key] = cached
            if done % 20 == 0 or done == len(batches):
                save_cache(cache)
                print(f"  进度 {done}/{len(batches)}", flush=True)
    save_cache(cache)

    # 应用结果
    modified = 0
    for i, e in enumerate(entries):
        r = cached.get(str(i))
        if r:
            new_rule = r.get("业务规则","")
            new_kal = r.get("业务口径","")
            old_rule = e.get("业务规则","") or ""
            old_kal = e.get("业务口径","") or ""
            if new_rule and new_rule != old_rule:
                e["业务规则"] = new_rule
                modified += 1
            if new_kal != old_kal:
                e["业务口径"] = new_kal
                modified += 1
    print(f"  修改: {modified} 处", flush=True)
    return modified

# ============ Phase 2: 反思检查 ============
def build_reflect_prompt(entries):
    items = []
    for i, e in enumerate(entries):
        cn = (e.get("标准中文名称","") or "")[:30]
        df = (e.get("业务定义","") or "")[:60]
        kal = (e.get("业务口径","") or "")[:60]
        rule = (e.get("业务规则","") or "")[:80]
        items.append(f'{i}. [{cn}] 定义:{df} | 口径:{kal} | 规则:{rule}')
    joined = "\n".join(items)
    return f"""你是银行业数据标准质检专家。请复查以下代码枚举类字段的"业务规则"和"业务口径"是否正确。

检查项：
1. 业务规则是否为标准"编码-名称;编码-名称"格式（行标国标引用除外）
2. 业务规则中的枚举值是否与字段名称和业务含义相关
3. 业务口径不应包含枚举值（如果包含，应移至业务规则并清空口径）
4. 双横线"--"、尾部多余符号（如"、"）是否已修复
5. 业务口径如果非空，应为字段来源/口径描述，不应是枚举值

只输出有问题的条目（没问题的不要输出）：
[{{"idx":0,"问题":"枚举值与字段含义不相关","修正业务规则":"1-男;2-女;0-未知","修正业务口径":""}}]

字段列表：
{joined}
"""

def phase_reflect(dictionary, cache):
    entries = dictionary["代码枚举类"]
    print(f"\n=== Phase 2: 反思检查 ({len(entries)}条) ===", flush=True)
    cache_key = "reflect"
    cached = cache.get(cache_key, {})
    todo = [(i, e) for i, e in enumerate(entries) if str(i) not in cached]
    print(f"  缓存 {len(cached)}, 新 {len(todo)}", flush=True)

    batches = [todo[j:j+BATCH] for j in range(0, len(todo), BATCH)]

    def _do(batch):
        batch_entries = [e for _, e in batch]
        try:
            text = call_llm(build_reflect_prompt(batch_entries), max_tokens=3000)
            arr = extract_json(text)
            local = {}
            for item in arr:
                idx = item.get("idx")
                if idx is not None and 0 <= idx < len(batch):
                    local[str(batch[idx][0])] = {
                        "修正业务规则": item.get("修正业务规则",""),
                        "修正业务口径": item.get("修正业务口径",""),
                    }
            return local
        except Exception as e:
            print(f"  反思批次异常: {str(e)[:80]}", flush=True)
            return {}

    done = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = {ex.submit(_do, b): b for b in batches}
        for fut in as_completed(futs):
            done += 1
            local = fut.result()
            cached.update(local)
            cache[cache_key] = cached
            if done % 20 == 0 or done == len(batches):
                save_cache(cache)
                print(f"  进度 {done}/{len(batches)}", flush=True)
    save_cache(cache)

    # 应用修正
    fixed = 0
    for i, e in enumerate(entries):
        r = cached.get(str(i))
        if r:
            if r.get("修正业务规则"):
                e["业务规则"] = r["修正业务规则"]
                fixed += 1
            if "修正业务口径" in r:
                e["业务口径"] = r["修正业务口径"]
                fixed += 1
    print(f"  反思修正: {fixed} 处", flush=True)
    return fixed

# ============ Phase 3: 导出 ============
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="4472C4")
WRAP = Alignment(wrap_text=True, vertical="top")

def style_sheet(ws, widths=None):
    for c in ws[1]:
        c.font = HDR_FONT; c.fill = HDR_FILL
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

def parse_enum_codes(rule):
    """从业务规则解析枚举编码明细"""
    if not rule:
        return []
    rule_str = str(rule).strip()
    # 行标国标引用，不解析
    if re.search(r'参考|参照|参见|《.*》|GB/?T|JR/T', rule_str):
        return []
    codes = []
    # 分割
    parts = re.split(r'[;；]', rule_str)
    for p in parts:
        p = p.strip().rstrip("、，,")
        if "-" in p:
            seg = p.split("-", 1)
            if len(seg) == 2:
                code = seg[0].strip()
                name = seg[1].strip()
                if code and name and code != "--":
                    codes.append((code, name))
    return codes

def phase_export(dictionary, domains):
    print(f"\n=== Phase 3: 导出 ===", flush=True)
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 全量字典 (合并6 sheet, 按标准编号排序)
    ws1 = wb.create_sheet("全量字典")
    ws1.append(DICT_COLS)
    all_entries = []
    for sheet_type in SHEET_TYPES:
        all_entries.extend(dictionary.get(sheet_type, []))
    # 按标准编号排序
    all_entries.sort(key=lambda e: e.get("标准编号","") or "")
    for e in all_entries:
        ws1.append([e.get(k,"") if e.get(k) is not None else "" for k in DICT_COLS])
    style_sheet(ws1, {"A":12,"B":22,"C":36,"D":22,"E":18,"F":10,"G":40,"H":20,"I":30,"J":20,"K":12,"L":12,"M":12})
    print(f"  全量字典: {len(all_entries)} 条", flush=True)

    # Sheet 2: 全量域
    ws2 = wb.create_sheet("全量域")
    h2 = ["域编号","域名称","域类型","标准所属类型","枚举值编号"]
    ws2.append(h2)
    for d in domains:
        ws2.append([d.get(k,"") if d.get(k) is not None else "" for k in h2])
    style_sheet(ws2, {"A":12,"B":30,"C":15,"D":12,"E":12})
    print(f"  全量域: {len(domains)} 域", flush=True)

    # Sheet 3: 枚举值表
    ws3 = wb.create_sheet("枚举值表")
    h3 = ["枚举值编号","枚举值名称","编码规则","域编号","域类型","枚举值数量"]
    ws3.append(h3)
    # Sheet 4: 枚举值编码明细
    ws4 = wb.create_sheet("枚举值编码明细")
    h4 = ["枚举值编号","枚举编码","枚举名称"]
    ws4.append(h4)

    enum_count = 0
    detail_count = 0
    for e in dictionary.get("代码枚举类", []):
        enm = e.get("枚举值编号","") or ""
        dom_name = e.get("域名称","") or ""
        cde = e.get("域编号","") or ""
        dtype = e.get("域类型","") or ""
        rule = e.get("业务规则","") or ""
        codes = parse_enum_codes(rule)
        ws3.append([enm, dom_name, rule, cde, dtype, len(codes)])
        enum_count += 1
        for code, name in codes:
            ws4.append([enm, code, name])
            detail_count += 1
    style_sheet(ws3, {"A":12,"B":30,"C":50,"D":12,"E":12,"F":10})
    style_sheet(ws4, {"A":12,"B":12,"C":20})
    print(f"  枚举值表: {enum_count} 条", flush=True)
    print(f"  枚举值编码明细: {detail_count} 条", flush=True)

    wb.save(OUT_PATH)
    print(f"\n导出: {OUT_PATH}", flush=True)

# ============ 主流程 ============
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--phase", default="all")
    args = ap.parse_args()
    cache = load_cache()

    # 载入或恢复状态
    state = load_state()
    if state and args.phase != "all":
        dictionary = state["dictionary"]
        domains = state["domains"]
        print(f"  从状态恢复: {sum(len(v) for v in dictionary.values())} 条", flush=True)
    else:
        dictionary = load_dictionary()
        domains = load_domains()

    if args.phase in ("all", "cleanup"):
        phase_cleanup(dictionary, cache)
        save_state({"dictionary": dictionary, "domains": domains})

    if args.phase in ("all", "reflect"):
        if args.phase == "reflect" and not state:
            print("  无状态，先运行 cleanup", flush=True)
            return
        phase_reflect(dictionary, cache)
        save_state({"dictionary": dictionary, "domains": domains})

    if args.phase in ("all", "export"):
        phase_export(dictionary, domains)

    print("\n完成!", flush=True)

if __name__ == "__main__":
    main()
