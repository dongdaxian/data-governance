"""Excel 读取/写入工具。

读取：将 Excel 解析为 FieldToMap 列表，按精确列名匹配（与质检模块一致）。
  输入 Excel 格式：第1行合并表头，第2行列名，第3行+数据。
写入：完整复制输入文件，追加落标结果列，保留原始格式。
  输出 Excel 格式：第1行合并表头（原始组 + "落标结果"组），第2行列名，第3行+数据。
"""

import shutil

import pandas as pd
from openpyxl import load_workbook

from standard_mapping.constants import (
    INPUT_COLUMNS,
    COL_MAPPING_RESULT,
    COL_SELECTED_STD_ID,
    COL_SELECTED_STD_NAME,
    COL_LLM_REASON,
    COL_CANDIDATES,
)
from standard_mapping.state import FieldToMap


def _safe_str(val) -> str:
    """安全转字符串，处理 NaN 和空值。"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    return s


def _match_columns(df: pd.DataFrame) -> dict[str, str]:
    """精确匹配 Excel 列名到内部字段名。

    申请单格式固定，按 INPUT_COLUMNS 中定义的精确列名直接查找。
    enum_values 支持多候选列名（优先匹配第一个找到的）。

    Returns:
        {"field_name": "实际列名", "field_type": "...", ...}
    """
    actual_cols = set(str(c) for c in df.columns)
    col_map = {}
    for internal_name, keywords in INPUT_COLUMNS.items():
        for kw in keywords:
            if kw in actual_cols:
                col_map[internal_name] = kw
                break
    return col_map


def read_excel(file_path: str) -> list[FieldToMap]:
    """读取 Excel 文件，返回 FieldToMap 列表。

    输入 Excel 第1行为合并表头，第2行为列名，从第3行开始为数据。
    """
    df = pd.read_excel(file_path, header=1)
    col_map = _match_columns(df)

    # 校验必填列是否全部匹配到
    required_keys = ["field_name", "field_type", "business_meaning", "data_example"]
    missing = [k for k in required_keys if k not in col_map]
    if missing:
        raise ValueError(
            f"输入Excel缺少必填列: {missing}，"
            f"请检查输入文件是否为质检输出格式。"
        )

    rows: list[FieldToMap] = []
    for idx, row in df.iterrows():
        rows.append(FieldToMap(
            index=int(idx),
            field_name=_safe_str(row.get(col_map["field_name"], "")),
            field_type=_safe_str(row.get(col_map["field_type"], "")),
            business_meaning=_safe_str(row.get(col_map["business_meaning"], "")),
            enum_values=_safe_str(row.get(col_map["enum_values"], "")) if "enum_values" in col_map else "",
            data_example=_safe_str(row.get(col_map["data_example"], "")),
            # 初始化结果字段
            candidates=[],
            candidate_fetch_error="",
            domain_check_details="",
            mapping_result="",
            selected_std_id="",
            selected_std_name="",
            llm_reason="",
        ))

    print(f"已读取 {len(rows)} 行数据")
    print(f"  列映射: {col_map}")
    return rows


def _format_candidates(candidates) -> str:
    """将候选标准格式化为明细文本（仅测试输出用）。"""
    if not candidates:
        return ""
    parts = []
    for c in candidates:
        sid = c["std_id"]
        sname = c["std_name"]
        stype = c["std_type"]
        dtype = c["domain_type"]
        ds = c.get("dense_score", 0.0)
        ss = c.get("sparse_score", 0.0)
        parts.append(f"{sid} {sname}({stype}/{dtype},dense={ds},sparse={ss})")
    return "; ".join(parts)


def write_excel(
    file_path: str,
    rows: list[FieldToMap],
    input_file: str,
    include_candidates: bool = False,
):
    """将落标结果写入 Excel。

    完整复制输入文件作为输出（保留全部原始列、数据、合并表头和格式），
    在末尾追加落标结果列，第1行新增"落标结果"合并表头组。
    代码枚举类等被过滤的行，结果列填空。
    """
    # 1. 完整复制输入文件作为输出文件
    shutil.copyfile(input_file, file_path)

    # 2. 打开输出文件，追加落标结果列
    wb = load_workbook(file_path)
    ws = wb.active
    max_col = ws.max_column

    # 定义新增列顺序
    result_cols = [COL_MAPPING_RESULT, COL_SELECTED_STD_ID, COL_SELECTED_STD_NAME, COL_LLM_REASON]
    if include_candidates:
        result_cols.append(COL_CANDIDATES)

    # 3. 第2行追加列名
    col_offset = max_col + 1
    for i, col_name in enumerate(result_cols):
        ws.cell(row=2, column=col_offset + i, value=col_name)

    # 4. 第1行追加"落标结果"合并表头
    start_col = col_offset
    end_col = col_offset + len(result_cols) - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    ws.cell(row=1, column=start_col, value="落标结果")

    # 5. 从第3行起逐行填入结果
    df = pd.read_excel(input_file, header=1)
    n = len(df)
    row_map = {r["index"]: r for r in rows}

    for idx in range(n):
        excel_row = 3 + idx
        r = row_map.get(idx)
        if r is not None:
            ws.cell(row=excel_row, column=col_offset, value=r["mapping_result"])
            ws.cell(row=excel_row, column=col_offset + 1, value=r["selected_std_id"])
            ws.cell(row=excel_row, column=col_offset + 2, value=r["selected_std_name"])
            ws.cell(row=excel_row, column=col_offset + 3, value=r["llm_reason"])
            if include_candidates:
                ws.cell(row=excel_row, column=col_offset + 4, value=_format_candidates(r["candidates"]))
        else:
            # 被过滤的行（如代码枚举类），结果列填空
            for i in range(len(result_cols)):
                ws.cell(row=excel_row, column=col_offset + i, value="")

    wb.save(file_path)

    print(f"结果已写入: {file_path}")

    # 打印统计
    sorted_rows = sorted(rows, key=lambda r: r["index"])
    total = len(sorted_rows)
    reuse = sum(1 for r in sorted_rows if r["mapping_result"] == "复用已有标准")
    reuse_ext = sum(1 for r in sorted_rows if r["mapping_result"] == "复用已有标准但扩展业务定义")
    new_std = sum(1 for r in sorted_rows if r["mapping_result"] == "新增标准")
    fetch_err = sum(1 for r in sorted_rows if "检索失败" in r["mapping_result"])
    print(f"  总计: {total} 行 | 复用已有标准: {reuse} | 复用但扩展: {reuse_ext} | 新增标准: {new_std} | 检索失败: {fetch_err}")
