# -*- coding: utf-8 -*-
"""生成 enum_standard_mapping 测试数据。

从质检测试数据（test/quality_check/output.xlsx）复制表头结构，
构造覆盖七分类场景的代码枚举类字段测试行。

用法：
  python test/enum_standard_mapping/gen_test_data.py
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
from openpyxl import load_workbook

QC_OUTPUT = os.path.join(os.path.dirname(__file__), "..", "quality_check", "output.xlsx")
OUT_DIR = os.path.dirname(__file__)


def main():
    # 读取质检输出（两层表头：第1行组名，第2行列名）
    df = pd.read_excel(QC_OUTPUT, header=1)
    cols = [c for c in df.columns if str(c) != "nan"]

    # 测试行：覆盖七分类 + 无枚举值/非枚举类型的边界行
    test_rows = [
        # 1. 复用已有标准：字段名与字典"性别代码"同名，码值 = 域CDE00002 子集
        {"中文字段名(必填)": "性别代码", "字段所属类型(必填)": "代码枚举类",
         "业务定义(必填)": "标识客户性别的代码", "格式化枚举值": "1-男;2-女;0-未知"},
        # 2. 补充域的枚举值后复用标准：字段名同名"渠道类型代码"，码值大量重复但有缺失
        {"中文字段名(必填)": "渠道类型代码", "字段所属类型(必填)": "代码枚举类",
         "业务定义(必填)": "办理业务时可通过的渠道类型", "格式化枚举值": "1-柜面;2-手机银行;3-网上银行;4-智能柜台"},
        # 3. 新增标准复用已有域：字段名不同（"客户性别代码"），码值与性别域完全重复
        {"中文字段名(必填)": "客户性别代码", "字段所属类型(必填)": "代码枚举类",
         "业务定义(必填)": "零售客户信息中的性别标识", "格式化枚举值": "1-男;2-女;0-未知"},
        # 4. 新增域并新增标准：码值与名称都检索不到
        {"中文字段名(必填)": "量子计算状态代码", "字段所属类型(必填)": "代码枚举类",
         "业务定义(必填)": "量子计算任务执行状态", "格式化枚举值": "A-量子叠加;B-量子坍缩;C-量子纠缠;D-量子退相干"},
        # 5. 非代码枚举类字段：应被过滤，不处理
        {"中文字段名(必填)": "客户名称", "字段所属类型(必填)": "文本类",
         "业务定义(必填)": "客户的姓名", "格式化枚举值": ""},
    ]

    # 构造与质检输出一致结构的 DataFrame（第2行列名 = cols）
    out_df = pd.DataFrame(test_rows, columns=cols)
    input_path = os.path.join(OUT_DIR, "input.xlsx")
    out_df.to_excel(input_path, index=False, startrow=1)

    # 第1行合并表头"基本信息"（to_excel startrow=1 已将列名写在第2行）
    wb = load_workbook(input_path)
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(row=1, column=1, value="基本信息")
    wb.save(input_path)
    print(f"测试输入已生成: {input_path}（{len(test_rows)} 行）")


if __name__ == "__main__":
    main()
